"""Excel Report Generator - Creates Excel reports using openpyxl."""

import structlog
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path
from enum import Enum

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side,
        NamedStyle, Color
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


logger = structlog.get_logger()


class ColorScheme(str, Enum):
    """统一配色方案."""

    # 主色调 - 蓝色系
    PRIMARY_DARK = "1F4E78"      # 深蓝 - 标题背景
    PRIMARY_MEDIUM = "4472C4"    # 中蓝 - 强调色
    PRIMARY_LIGHT = "D9E1F2"     # 浅蓝 - 表头背景

    # 辅助色
    ACCENT_GRAY = "F2F2F2"       # 浅灰 - 交替行背景
    BORDER_COLOR = "B7B7B7"      # 边框灰

    # 状态色
    SUCCESS = "70AD47"           # 绿色 - 良好
    WARNING = "FFC000"           # 黄色 - 警告
    DANGER = "FF0000"            # 红色 - 危险
    INFO = "5B9BD5"              # 信息蓝


class ExcelStyles:
    """统一样式管理器."""

    # 标题样式
    TITLE_FONT = Font(size=16, bold=True, color="FFFFFF")
    TITLE_FILL = PatternFill(start_color=ColorScheme.PRIMARY_DARK,
                             end_color=ColorScheme.PRIMARY_DARK,
                             fill_type="solid")
    TITLE_ALIGNMENT = Alignment(horizontal='center', vertical='center')

    # 副标题样式
    SUBTITLE_FONT = Font(size=12, bold=True, color=ColorScheme.PRIMARY_DARK)
    SUBTITLE_FILL = PatternFill(start_color=ColorScheme.PRIMARY_LIGHT,
                                end_color=ColorScheme.PRIMARY_LIGHT,
                                fill_type="solid")

    # 表头样式
    HEADER_FONT = Font(size=11, bold=True, color=ColorScheme.PRIMARY_DARK)
    HEADER_FILL = PatternFill(start_color=ColorScheme.PRIMARY_LIGHT,
                             end_color=ColorScheme.PRIMARY_LIGHT,
                             fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)

    # 数据单元格边框
    THIN_BORDER = Border(
        left=Side(style='thin', color=ColorScheme.BORDER_COLOR),
        right=Side(style='thin', color=ColorScheme.BORDER_COLOR),
        top=Side(style='thin', color=ColorScheme.BORDER_COLOR),
        bottom=Side(style='thin', color=ColorScheme.BORDER_COLOR),
    )

    # 对齐方式
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

    # 数值格式
    PERCENT_FORMAT = '0.0%'
    NUMBER_FORMAT = '#,##0.00'
    INTEGER_FORMAT = '#,##0'
    DATE_FORMAT = 'yyyy-mm-dd hh:mm'
    DATE_SHORT_FORMAT = 'yyyy-mm-dd'


class RiskColors:
    """风险等级颜色映射."""

    COLORS = {
        # 英文键（原始数据）
        "critical": ColorScheme.DANGER,
        "high": "FF6B6B",
        "medium": ColorScheme.WARNING,
        "low": ColorScheme.SUCCESS,
        # 中文键（标准化后）
        "危急": ColorScheme.DANGER,
        "高": "FF6B6B",
        "中": ColorScheme.WARNING,
        "低": ColorScheme.SUCCESS,
        # 默认无颜色
        "": None,
    }

    @classmethod
    def get_fill(cls, risk_level: str) -> PatternFill:
        """获取风险等级对应的填充色."""
        if not isinstance(risk_level, str):
            risk_level = ""

        # 先尝试直接匹配，再尝试小写匹配
        color = cls.COLORS.get(risk_level) or cls.COLORS.get(risk_level.lower())

        # 如果没有匹配到颜色，返回无填充
        if color is None:
            return PatternFill(fill_type="none")

        return PatternFill(start_color=color, end_color=color, fill_type="solid")


class HealthColors:
    """健康评分颜色映射."""

    @classmethod
    def get_color(cls, score: float) -> str:
        """根据分数获取颜色."""
        if score >= 80:
            return ColorScheme.SUCCESS
        elif score >= 60:
            return ColorScheme.WARNING
        elif score >= 40:
            return "#FF9900"
        else:
            return ColorScheme.DANGER


# Column definitions for each sheet
# Format: (field_key, header_text, data_type, width)
# data_type: 'text', 'number', 'percent', 'date'
COLUMNS = {
    "clusters": [
        ("name", "名称", "text", 25),
        ("datacenter", "数据中心", "text", 20),
        ("total_cpu", "CPU (MHz)", "number", 14),
        ("total_memory_gb", "内存 (GB)", "number", 14),
        ("total_storage_gb", "存储总量 (GB)", "number", 14),
        ("used_storage_gb", "已用存储 (GB)", "number", 14),
        ("num_hosts", "主机数", "integer", 10),
        ("num_vms", "虚拟机数", "integer", 12),
    ],
    "hosts": [
        ("name", "名称", "text", 25),
        ("datacenter", "数据中心", "text", 18),
        ("ip_address", "IP地址", "text", 15),
        ("cpu_cores", "CPU核数", "integer", 10),
        ("cpu_mhz", "频率 (MHz)", "number", 12),
        ("memory_gb", "内存 (GB)", "number", 12),
        ("num_vms", "虚拟机数", "integer", 10),
        ("power_state", "电源状态", "text", 12),
        ("overall_status", "状态", "text", 10),
    ],
    # 与前端"虚拟机列表"Tab 完全对齐：虚拟机名称、CPU、内存、磁盘使用、状态、数据中心、主机IP
    "vms": [
        ("name", "虚拟机名称", "text", 30),
        ("cpu_count", "CPU (核)", "integer", 10),
        ("memory_gb", "内存 (GB)", "number", 12),
        ("disk_usage_gb", "磁盘使用 (GB)", "number", 14),
        ("power_state", "状态", "text", 12),
        ("datacenter", "数据中心", "text", 18),
        ("host_ip", "主机IP", "text", 15),
    ],
    # 与前端"闲置检测"Tab 完全对齐：虚拟机→状态→集群→主机IP→闲置类型→风险等级→闲置天数→最后活动→置信度→优化建议
    "idle": [
        ("vmName", "虚拟机", "text", 30),
        ("isIdle", "状态", "text", 8),
        ("cluster", "集群", "text", 18),
        ("hostIp", "主机IP", "text", 15),
        ("idleType", "闲置类型", "text", 12),
        ("riskLevel", "风险等级", "text", 12),
        ("daysInactive", "闲置天数", "integer", 10),
        ("lastActivityTime", "最后活动", "date", 18),
        ("confidence", "置信度(%)", "number", 10),
        ("recommendation", "优化建议", "text", 45),
    ],
    # 与前端"资源优化"Tab 完全对齐
    "rightsize": [
        ("vmName", "虚拟机", "text", 30),
        ("cluster", "集群", "text", 18),
        ("hostIp", "主机IP", "text", 15),
        ("currentCpu", "当前CPU(核)", "integer", 12),
        ("currentMemoryGb", "当前内存(GB)", "number", 14),
        ("recommendedCpu", "推荐CPU(核)", "integer", 12),
        ("recommendedMemoryGb", "推荐内存(GB)", "number", 14),
        ("cpuP95", "CPU P95(%)", "number", 12),
        ("cpuP90", "CPU P90(%)", "number", 12),
        ("cpuAvg", "CPU均值(%)", "number", 12),
        ("memoryP95", "内存P95(%)", "number", 12),
        ("memoryAvg", "内存均值(%)", "number", 12),
        ("mismatchType", "错配类型", "text", 20),
        ("cpuWasteRatio", "CPU浪费比例(%)", "number", 14),
        ("memWasteRatio", "内存浪费比例(%)", "number", 14),
        ("cpuAdjustmentType", "CPU调整方向", "text", 15),
        ("memAdjustmentType", "内存调整方向", "text", 15),
        ("confidence", "置信度(%)", "number", 10),
        ("reason", "判断依据", "text", 55),
    ],
    # 与前端"潮汐检测"Tab 完全对齐
    "tidal": [
        ("vmName", "虚拟机", "text", 30),
        ("cluster", "集群", "text", 18),
        ("hostIp", "主机IP", "text", 15),
        ("tidalGranularity", "潮汐粒度", "text", 12),
        ("volatilityLevel", "波动程度", "text", 12),
        ("coefficientOfVariation", "CV变异系数", "number", 12),
        ("peakValleyRatio", "峰谷比", "number", 10),
        ("recommendedOffHours_description", "建议关机时段", "text", 28),
        ("reason", "分析依据", "text", 55),
    ],
    # 分析结果 - 可释放主机列表（与前端"分析结果"Tab 对齐）
    "freeable_hosts": [
        ("hostName", "主机名", "text", 30),
        ("hostIp", "主机IP", "text", 15),
        ("cpuCores", "CPU核数", "integer", 12),
        ("memoryGb", "内存 (GB)", "number", 14),
        ("storageGb", "存储 (GB)", "number", 14),
        ("currentVmCount", "当前VM数", "integer", 12),
        ("reason", "原因", "text", 60),
    ],
}


# 闲置类型中文映射
IDLE_TYPE_MAP = {
    "powered_off": "已关机",
    "idle_powered_on": "开机闲置",
    "low_activity": "低活跃",
}


# 风险等级中文映射
RISK_LEVEL_MAP = {
    "critical": "危急",
    "high": "高",
    "medium": "中",
    "low": "低",
}


# 使用模式中文映射
USAGE_PATTERN_MAP = {
    "stable": "稳定",
    "burst": "突发",
    "tidal": "潮汐",
    "unknown": "未知",
}


# 错配类型中文映射
MISMATCH_TYPE_MAP = {
    "cpu_rich_memory_poor": "CPU富足/内存不足",
    "cpu_poor_memory_rich": "CPU不足/内存富足",
    "both_underutilized": "均利用不足",
    "both_overutilized": "均利用过高",
    "balanced": "配比合理",
}


class ExcelReportGenerator:
    """Generates Excel reports using openpyxl."""

    def __init__(self) -> None:
        """Initialize Excel report generator."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel report generation")

    def generate(
        self,
        data: Dict[str, Any],
        output_path: str,
    ) -> str:
        """Generate Excel report.

        Args:
            data: Report data from ReportBuilder
            output_path: Output file path

        Returns:
            Generated file path
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets in specific order
        self._create_summary_sheet(wb, data)
        self._create_analysis_result_sheet(wb, data)
        self._create_clusters_sheet(wb, data)
        self._create_hosts_sheet(wb, data)
        self._create_vms_sheet(wb, data)
        self._create_idle_sheet(wb, data)
        self._create_rightsize_sheet(wb, data)
        self._create_tidal_sheet(wb, data)
        self._create_health_sheet(wb, data)

        # Set active sheet to summary
        wb.active = 0

        # Save workbook
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info("excel_report_generated", path=str(output_path))
        return str(output_path)

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create summary sheet with modern dashboard style layout."""
        ws = wb.create_sheet("概览", 0)

        # 列宽设置 - 第一列设为第二列的70%
        ws.column_dimensions["A"].width = 28   # 40 * 0.7
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 3

        current_row = 1

        # ==================== 标题区域 ====================
        ws.merge_cells(f"A{current_row}:E{current_row}")
        ws.row_dimensions[current_row].height = 35
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = "云平台资源评估报告"
        title_cell.font = ExcelStyles.TITLE_FONT
        title_cell.fill = ExcelStyles.TITLE_FILL
        title_cell.alignment = ExcelStyles.TITLE_ALIGNMENT
        current_row += 2

        # ==================== 核心指标卡片区 ====================
        # 区域标题
        self._write_section_header(ws, current_row, "📊 核心指标", 1, 5)
        current_row += 1

        summary = data.get("summary", {})

        # 第一行：4个指标卡片
        metrics_row_1 = [
            ("集群总数", summary.get("total_clusters", 0), "个"),
            ("主机总数", summary.get("total_hosts", 0), "台"),
            ("选择VM", summary.get("total_vms", 0), "个"),
            ("开机VM", summary.get("powered_on_vms", 0), "个"),
        ]

        for i, (label, value, unit) in enumerate(metrics_row_1):
            col = 1 + i * 1  # A, B, C, D
            if col <= 4:
                self._write_metric_card(ws, current_row, col, label, value, unit, i)

        current_row += 3

        # 第二行：关机VM卡片（与开机VM同列对齐，放D列）
        powered_off = summary.get("powered_off_vms", 0)
        self._write_metric_card(ws, current_row, 4, "关机VM", powered_off, "个", 1, accent=True)
        current_row += 3

        # ==================== 资源总量区域 ====================
        self._write_section_header(ws, current_row, "💾 资源总量", 1, 5)
        current_row += 1

        # 使用表格样式显示资源（去掉进度条）
        total_cpu = summary.get("total_cpu_mhz") or 0
        total_memory = summary.get("total_memory_gb") or 0
        total_storage = summary.get("total_storage_gb") or 0
        used_storage = summary.get("used_storage_gb") or 0

        self._write_resource_item(ws, current_row, "总CPU频率", f"{total_cpu:,} MHz", ColorScheme.PRIMARY_MEDIUM)
        current_row += 2

        self._write_resource_item(ws, current_row, "总内存容量", f"{total_memory:,.2f} GB", ColorScheme.PRIMARY_MEDIUM)
        current_row += 2

        self._write_resource_item(ws, current_row, "存储总量", f"{total_storage:,.2f} GB", ColorScheme.PRIMARY_MEDIUM)
        current_row += 2

        self._write_resource_item(ws, current_row, "已用存储", f"{used_storage:,.2f} GB", ColorScheme.PRIMARY_MEDIUM)
        current_row += 3

        # ==================== 优化建议区域 ====================
        self._write_section_header(ws, current_row, "💡 优化建议", 1, 5)
        current_row += 1

        self._write_optimization_cards(ws, current_row, data)
        current_row += 8

        # ==================== 底部信息 ====================
        task = data.get("task", {})
        connection = data.get("connection", {})
        generated_at = self._format_datetime(data.get("generated_at"))

        # 左侧：任务信息
        ws.cell(row=current_row, column=1, value="📋 任务信息").font = Font(bold=True, size=10, color="666666")
        ws.cell(row=current_row + 1, column=1, value=f"任务名称: {task.get('name') or '-'}").font = Font(size=9)

        platform = connection.get("platform")
        platform_display = platform.upper() if platform else "-"
        ws.cell(row=current_row + 2, column=1, value=f"连接平台: {platform_display}").font = Font(size=9)
        ws.cell(row=current_row + 3, column=1, value=f"生成时间: {generated_at}").font = Font(size=9)

        # 右侧：状态
        ws.cell(row=current_row, column=4, value="✓ 状态").font = Font(bold=True, size=10, color="666666")
        ws.cell(row=current_row + 1, column=4, value=f"任务状态: {self._format_status(task.get('status', '-'))}").font = Font(size=9)
        ws.cell(row=current_row + 2, column=4, value=f"报告类型: Excel 报告").font = Font(size=9)

    def _write_metric_card(
        self,
        ws,
        row: int,
        col: int,
        label: str,
        value: int,
        unit: str,
        index: int,
        accent: bool = False,
    ) -> None:
        """Write a metric card with number visualization.

        Args:
            ws: Worksheet
            row: Start row
            col: Column (1=A, 2=B, etc.)
            label: Metric label
            value: Metric value
            unit: Unit string
            index: Card index for color variation
            accent: Whether to use accent color
        """
        # 处理空值
        safe_value = 0 if value is None else value

        # 定义卡片颜色
        card_colors = ["D9E1F2", "E2EFDA", "FCE4D6", "E7E6E6"]
        fill_color = card_colors[index % len(card_colors)] if not accent else "F4B084"

        # 卡片背景
        card_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # 标签行
        label_cell = ws.cell(row=row, column=col)
        label_cell.value = label
        label_cell.font = Font(size=10, bold=True, color="505050")
        label_cell.fill = card_fill
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        label_cell.border = ExcelStyles.THIN_BORDER

        # 数值行
        value_cell = ws.cell(row=row + 1, column=col)
        value_cell.value = f"{safe_value:,} {unit}"
        value_cell.font = Font(size=20, bold=True, color="1F4E78" if not accent else "C00000")
        value_cell.fill = card_fill
        value_cell.alignment = Alignment(horizontal='center', vertical='center')
        value_cell.border = ExcelStyles.THIN_BORDER

    def _write_resource_item(
        self,
        ws,
        row: int,
        label: str,
        value: str,
        color: str,
    ) -> None:
        """Write a simple resource item display.

        Args:
            ws: Worksheet
            row: Row number
            label: Label text
            value: Value to display (formatted string)
            color: Accent color
        """
        # 标签
        label_cell = ws.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = Font(bold=True, size=11)

        # 数值（带颜色突出）
        value_cell = ws.cell(row=row, column=2)
        value_cell.value = value
        value_cell.font = Font(size=14, bold=True, color=color)
        value_cell.alignment = Alignment(horizontal='left')

    def _write_optimization_cards(
        self,
        ws,
        start_row: int,
        data: Dict[str, Any],
    ) -> None:
        """Write optimization recommendation cards.

        Args:
            ws: Worksheet
            start_row: Start row
            data: Report data
        """
        analysis = data.get("analysis", {})
        idle = analysis.get("idle", [])
        resource = analysis.get("resource", {})
        health = analysis.get("health", {})

        # 闲置VM摘要
        idle_count = len(idle)
        critical_idle = sum(1 for i in idle if i.get("riskLevel") == "critical")

        # 第一行：闲置VM + 可优化配置（并排）
        self._write_opt_card(ws, start_row, 1, "⚠️", "闲置虚拟机",
                           f"{idle_count} 个", f"其中 {critical_idle} 个需要紧急处理",
                           ColorScheme.WARNING)

        rightsize = resource.get("resourceOptimization", [])
        downsize_count = sum(1 for r in rightsize if (r.get("cpuAdjustmentType") or "").startswith("down") or (r.get("memAdjustmentType") or "").startswith("down"))

        self._write_opt_card(ws, start_row, 3, "📉", "可优化配置",
                           f"{len(rightsize)} 个", f"建议降配 {downsize_count} 个",
                           ColorScheme.INFO)

        start_row += 4

        # 第二行：健康评分（带蓝底标题）
        self._write_section_header(ws, start_row, "💚 平台健康评分", 1, 5)
        start_row += 1
        if health:
            overall_score = health.get("overallScore", 0)
            health_grade = health.get("grade", "unknown")
            grade_text = self._format_grade(health_grade)

            score_color = HealthColors.get_color(overall_score)
            self._write_opt_card(ws, start_row, 1, "💯", f"{overall_score} 分",
                               f"评级: {grade_text}", "",
                               score_color, is_color_code=True)

    def _write_opt_card(
        self,
        ws,
        row: int,
        col: int,
        icon: str,
        title: str,
        main_value: str,
        sub_value: str,
        color: str,
        is_color_code: bool = False,
    ) -> None:
        """Write an optimization card.

        Args:
            ws: Worksheet
            row: Row number
            col: Column number (1=A, 3=C)
            icon: Icon emoji
            title: Card title
            main_value: Main value to display
            sub_value: Subtitle/value
            color: Accent color (hex or ColorScheme enum)
            is_color_code: True if color is a hex code string
        """
        # 解析颜色
        if is_color_code:
            color_hex = color
        elif isinstance(color, str) and not color.startswith("#"):
            # ColorScheme enum value
            color_hex = ColorScheme[color].value if hasattr(ColorScheme, color) else color
        else:
            color_hex = color

        # 计算列范围（每个卡片占2列）
        start_col = col
        end_col = col + 1

        # 标题行
        title_cell = ws.cell(row=row, column=start_col)
        title_cell.value = f"{icon} {title}"
        title_cell.font = Font(bold=True, size=11, color=color_hex)
        ws.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")

        # 主数值行
        main_cell = ws.cell(row=row + 1, column=start_col)
        main_cell.value = main_value
        main_cell.font = Font(size=16, bold=True, color=color_hex)
        ws.merge_cells(f"{get_column_letter(start_col)}{row + 1}:{get_column_letter(end_col)}{row + 1}")

        # 副数值行
        sub_cell = ws.cell(row=row + 2, column=start_col)
        sub_cell.value = sub_value
        sub_cell.font = Font(size=10, color="666666")
        ws.merge_cells(f"{get_column_letter(start_col)}{row + 2}:{get_column_letter(end_col)}{row + 2}")

    def _create_analysis_result_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create analysis result sheet - 综合优化建议 + 可释放主机（对应前端"分析结果"Tab）."""
        ws = wb.create_sheet("分析结果")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 35

        analysis = data.get("analysis", {})
        freeability = analysis.get("freeability") or {}
        recommendation = freeability.get("recommendation") or {}
        post_opt = recommendation.get("postOptimization") or {}
        resource_opt = recommendation.get("resourceOptimization") or {}
        idle_det = recommendation.get("idleDetection") or {}

        current_row = 1

        # ==================== 优化措施 ====================
        self._write_section_header(ws, current_row, "优化措施", 1, 4)
        current_row += 1

        measures = []
        if resource_opt.get("enabled") and resource_opt.get("vmCount", 0) > 0:
            measures.append(
                f"资源优化：对 {resource_opt['vmCount']} 台VM进行Right Size，"
                f"释放 {resource_opt.get('freedCpuCores', 0)} 核CPU / {resource_opt.get('freedMemoryGb', 0):.1f} GB内存"
            )
        if idle_det.get("enabled") and idle_det.get("vmCount", 0) > 0:
            powered_on = idle_det.get("poweredOnCount", 0)
            powered_off = idle_det.get("poweredOffCount", 0)
            type_detail = ""
            if powered_on > 0 and powered_off > 0:
                type_detail = f"（开机闲置 {powered_on} 台、关机 {powered_off} 台）"
            elif powered_on > 0:
                type_detail = f"（开机闲置 {powered_on} 台）"
            elif powered_off > 0:
                type_detail = f"（关机 {powered_off} 台）"
            if powered_on > 0:
                savings_text = (
                    f"开机闲置VM关闭后释放 {idle_det.get('freedCpuCores', 0)} 核CPU / "
                    f"{idle_det.get('freedMemoryGb', 0):.1f} GB内存"
                )
                disk_on = idle_det.get('freedDiskPoweredOn', 0)
                if disk_on > 0:
                    savings_text += f" / {disk_on:.1f} GB磁盘"
                savings_text += "。"
            else:
                savings_text = ""
            if powered_off > 0:
                disk_off = idle_det.get('freedDiskPoweredOff', 0)
                if disk_off > 0:
                    savings_text += f"关机VM清除后释放 {disk_off:.1f} GB磁盘。"
            measures.append(
                f"闲置检测：检测到 {idle_det['vmCount']} 台闲置VM{type_detail}，{savings_text}"
            )

        if measures:
            for measure in measures:
                ws.cell(row=current_row, column=1, value=f"• {measure}").alignment = ExcelStyles.ALIGN_LEFT
                ws.merge_cells(f"A{current_row}:D{current_row}")
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="暂无优化措施").font = Font(italic=True, color="999999")
            current_row += 1

        current_row += 1

        # ==================== 集群资源需求 ====================
        self._write_section_header(ws, current_row, "集群资源需求（优化后）", 1, 4)
        current_row += 1

        req_items = [
            ("所需CPU核数", post_opt.get("neededCpuCores", "-")),
            ("所需内存(GB)", post_opt.get("neededMemoryGb", "-")),
            ("所需存储(GB)", post_opt.get("neededStorageGb", "-")),
            ("保留主机数", freeability.get("retainedHostCount", "-")),
        ]
        for i, (label, value) in enumerate(req_items):
            col = 1 + (i % 2) * 2
            row = current_row + i // 2
            ws.cell(row=row, column=col, value=label).font = Font(bold=True)
            ws.cell(row=row, column=col + 1, value=value)
        current_row += (len(req_items) + 1) // 2 + 1

        # ==================== 可释放主机 ====================
        self._write_section_header(ws, current_row, "可下线主机列表", 1, 4)
        current_row += 1

        freeable_hosts = freeability.get("freeableHosts") or []
        if freeable_hosts:
            self._write_table(ws, freeable_hosts, COLUMNS["freeable_hosts"],
                              add_row_colors=True, start_row=current_row)
            current_row += len(freeable_hosts) + 2
        else:
            ws.cell(row=current_row, column=1, value="当前优化方案下无可下线主机").font = Font(italic=True, color="999999")
            current_row += 2

        # ==================== 负载健康 ====================
        self._write_section_header(ws, current_row, "优化后负载预估", 1, 4)
        current_row += 1

        load_items = [
            ("CPU负载", post_opt.get("cpuLoadPercent")),
            ("内存负载", post_opt.get("memoryLoadPercent")),
            ("存储负载", post_opt.get("storageLoadPercent")),
            ("健康状态", post_opt.get("healthLabel")),
        ]
        for label, val in load_items:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            if isinstance(val, (int, float)):
                display = f"{val:.1f}%"
            else:
                display = str(val) if val is not None else "-"
            ws.cell(row=current_row, column=2, value=display)
            current_row += 1

    def _create_clusters_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create clusters sheet with enhanced formatting."""
        ws = wb.create_sheet("集群")
        resources = data.get("resources", {})
        clusters = resources.get("clusters", [])

        if not clusters:
            self._write_no_data(ws)
            return

        self._write_table(ws, clusters, COLUMNS["clusters"], add_row_colors=True)

    def _create_hosts_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create hosts sheet with enhanced formatting."""
        ws = wb.create_sheet("主机")
        resources = data.get("resources", {})
        hosts = resources.get("hosts", [])

        if not hosts:
            self._write_no_data(ws)
            return

        self._write_table(ws, hosts, COLUMNS["hosts"], add_row_colors=True)

    def _create_vms_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create VMs sheet with enhanced formatting."""
        ws = wb.create_sheet("虚拟机")
        resources = data.get("resources", {})
        vms = resources.get("vms", [])

        if not vms:
            self._write_no_data(ws)
            return

        self._write_table(ws, vms, COLUMNS["vms"], add_row_colors=True)

    def _create_idle_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create idle detection sheet with enhanced formatting."""
        ws = wb.create_sheet("闲置检测")
        analysis = data.get("analysis", {})
        idle_results = analysis.get("idle", [])

        if not idle_results:
            self._write_no_data(ws)
            return

        # 标准化数据，确保所有字段都存在（与前端"闲置检测"Tab 列顺序对齐）
        normalized_data = []
        for item in idle_results:
            is_idle = item.get("isIdle", False)
            normalized_item = {
                "vmName": item.get("vmName", ""),
                "isIdle": "闲置" if is_idle else "正常",
                "cluster": item.get("cluster", ""),
                "hostIp": item.get("hostIp", ""),
                "idleType": IDLE_TYPE_MAP.get(item.get("idleType", ""), item.get("idleType", "")),
                "riskLevel": RISK_LEVEL_MAP.get(item.get("riskLevel", ""), item.get("riskLevel", "")),
                "daysInactive": item.get("daysInactive"),
                "lastActivityTime": item.get("lastActivityTime"),
                "confidence": item.get("confidence", 0),
                "recommendation": item.get("recommendation", ""),
            }
            normalized_data.append(normalized_item)

        # 写入表格，带风险等级颜色
        self._write_table(
            ws,
            normalized_data,
            COLUMNS["idle"],
            add_row_colors=True,
            color_column="riskLevel",
            color_mapper=RiskColors.get_fill
        )

    def _create_rightsize_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create resource optimization sheet."""
        ws = wb.create_sheet("资源优化")
        analysis = data.get("analysis", {})
        resource = analysis.get("resource", {})
        rightsize = resource.get("resourceOptimization", [])

        if not rightsize:
            self._write_no_data(ws)
            return

        normalized_data = []
        for item in rightsize:
            normalized_data.append({
                "vmName": item.get("vmName", ""),
                "cluster": item.get("cluster", ""),
                "hostIp": item.get("hostIp", ""),
                "currentCpu": item.get("currentCpu", 0),
                "currentMemoryGb": item.get("currentMemoryGb", 0),
                "recommendedCpu": item.get("recommendedCpu", 0),
                "recommendedMemoryGb": item.get("recommendedMemoryGb", 0),
                "cpuP95": item.get("cpuP95", 0),
                "cpuP90": item.get("cpuP90", 0),
                "cpuAvg": item.get("cpuAvg", 0),
                "memoryP95": item.get("memoryP95", 0),
                "memoryAvg": item.get("memoryAvg", 0),
                "mismatchType": MISMATCH_TYPE_MAP.get(item.get("mismatchType", ""), item.get("mismatchType", "")),
                "cpuWasteRatio": item.get("cpuWasteRatio", 0),
                "memWasteRatio": item.get("memWasteRatio", 0),
                "cpuAdjustmentType": self._format_adjustment_type(item.get("cpuAdjustmentType") or ""),
                "memAdjustmentType": self._format_adjustment_type(item.get("memAdjustmentType") or ""),
                "confidence": item.get("confidence", 0),
                "reason": item.get("reason", ""),
            })

        self._write_table(ws, normalized_data, COLUMNS["rightsize"], add_row_colors=True)

    def _create_tidal_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create tidal detection sheet."""
        ws = wb.create_sheet("潮汐检测")
        analysis = data.get("analysis", {})
        resource = analysis.get("resource", {})
        tidal_items = resource.get("tidal", [])

        if not tidal_items:
            self._write_no_data(ws)
            return

        granularity_map = {"daily": "日粒度", "weekly": "周粒度", "monthly": "月粒度"}
        volatility_map = {"high": "高", "moderate": "中", "low": "低"}

        normalized_data = []
        for item in tidal_items:
            off_hours = item.get("recommendedOffHours") or {}
            normalized_data.append({
                "vmName": item.get("vmName", ""),
                "cluster": item.get("cluster", ""),
                "hostIp": item.get("hostIp", ""),
                "tidalGranularity": granularity_map.get(item.get("tidalGranularity", ""), item.get("tidalGranularity", "")),
                "volatilityLevel": volatility_map.get(item.get("volatilityLevel", ""), item.get("volatilityLevel", "")),
                "coefficientOfVariation": item.get("coefficientOfVariation", 0),
                "peakValleyRatio": item.get("peakValleyRatio", 0),
                "recommendedOffHours_description": off_hours.get("description", ""),
                "reason": item.get("reason", ""),
            })

        self._write_table(ws, normalized_data, COLUMNS["tidal"], add_row_colors=True)

    def _create_health_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create health score sheet with enhanced formatting."""
        ws = wb.create_sheet("健康评分")

        # 设置列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 40

        analysis = data.get("analysis", {})
        health = analysis.get("health", {})

        if not health:
            self._write_no_data(ws, message="无健康评分数据")
            return

        current_row = 1

        # ==================== 整体评分 ====================
        self._write_section_header(ws, current_row, "整体评分", 1, 3)
        current_row += 1

        overall_score = health.get("overallScore", 0)
        grade = health.get("grade", "unknown")

        # 评分大字显示
        score_cell = ws.cell(row=current_row, column=2)
        score_cell.value = overall_score
        score_cell.font = Font(size=36, bold=True, color=HealthColors.get_color(overall_score))
        score_cell.alignment = ExcelStyles.ALIGN_CENTER

        # 等级显示
        grade_text = self._format_grade(grade)
        grade_cell = ws.cell(row=current_row + 1, column=2)
        grade_cell.value = grade_text
        grade_cell.font = Font(size=14, bold=True)
        grade_cell.alignment = ExcelStyles.ALIGN_CENTER

        current_row += 3

        # ==================== 维度评分 ====================
        self._write_section_header(ws, current_row, "维度评分", 1, 3)
        current_row += 1

        # 表头
        headers = ["维度", "评分", "评级"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
            cell.border = ExcelStyles.THIN_BORDER
            cell.alignment = ExcelStyles.ALIGN_CENTER
        current_row += 1

        # 维度数据
        dimensions = [
            ("资源均衡度", "balanceScore"),
            ("超配风险", "overcommitScore"),
            ("热点集中度", "hotspotScore"),
        ]

        for label, key in dimensions:
            score = health.get(key, 0)
            ws.cell(row=current_row, column=1, value=label).border = ExcelStyles.THIN_BORDER
            ws.cell(row=current_row, column=1).alignment = ExcelStyles.ALIGN_LEFT

            score_cell = ws.cell(row=current_row, column=2, value=score)
            score_cell.border = ExcelStyles.THIN_BORDER
            score_cell.alignment = ExcelStyles.ALIGN_CENTER
            score_cell.font = Font(color=HealthColors.get_color(score))

            grade_cell = ws.cell(row=current_row, column=3, value=self._get_score_grade(score))
            grade_cell.border = ExcelStyles.THIN_BORDER
            grade_cell.alignment = ExcelStyles.ALIGN_CENTER

            current_row += 1

        current_row += 1

        # ==================== 平台概览 ====================
        cluster_count = health.get("clusterCount", 0)
        host_count = health.get("hostCount", 0)
        vm_count = health.get("vmCount", 0)

        self._write_section_header(ws, current_row, "平台概览", 1, 3)
        current_row += 1

        overview_data = [
            ("集群数量", cluster_count),
            ("主机数量", host_count),
            ("虚拟机数量", vm_count),
        ]

        for label, value in overview_data:
            ws.cell(row=current_row, column=1, value=label).alignment = ExcelStyles.ALIGN_LEFT
            ws.cell(row=current_row, column=2, value=value).alignment = ExcelStyles.ALIGN_CENTER
            current_row += 1

        current_row += 1

        # ==================== 发现的问题 ====================
        findings = health.get("findings", [])
        if findings:
            self._write_section_header(ws, current_row, "发现的问题", 1, 3)
            current_row += 1

            category_map = {
                "overcommit": "超配风险",
                "balance": "资源均衡",
                "hotspot": "热点集中",
            }
            severity_map = {
                "critical": "危急",
                "high": "高",
                "medium": "中",
                "low": "低",
            }

            for finding in findings:
                category = finding.get("category", "")
                severity = finding.get("severity", "")
                title = category_map.get(category, category) + (f"（{severity_map.get(severity, severity)}）" if severity else "")
                description = finding.get("description", "")

                title_cell = ws.cell(row=current_row, column=1)
                title_cell.value = "> " + title
                title_cell.font = Font(bold=True, color=ColorScheme.PRIMARY_MEDIUM)
                ws.merge_cells(f"A{current_row}:C{current_row}")
                current_row += 1

                desc_cell = ws.cell(row=current_row, column=1)
                desc_cell.value = "  " + description
                ws.merge_cells(f"A{current_row}:C{current_row}")
                current_row += 1

    # ==================== 辅助方法 ====================

    def _write_table(
        self,
        ws,
        data: List[Dict[str, Any]],
        columns: List[tuple],
        add_row_colors: bool = False,
        color_column: Optional[str] = None,
        color_mapper: Optional[callable] = None,
        percent_columns: Optional[List[str]] = None,
        start_row: int = 1,
    ) -> None:
        """Write data table to worksheet with enhanced formatting.

        Args:
            ws: Worksheet object
            data: List of data dicts
            columns: List of (field, header, type, width) tuples
            add_row_colors: Whether to add alternating row colors
            color_column: Column name to apply custom color
            color_mapper: Function to get fill color based on column value
            percent_columns: List of columns to format as percentage
        """
        if percent_columns is None:
            percent_columns = []

        if not data:
            self._write_no_data(ws, row=start_row)
            return

        header_row = start_row
        # Write headers
        for col_idx, (_, header, data_type, width) in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
            cell.border = ExcelStyles.THIN_BORDER
            cell.alignment = ExcelStyles.HEADER_ALIGNMENT
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row only when table starts at row 1
        if start_row == 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # Write data
        for row_idx, item in enumerate(data, header_row + 1):
            # Determine row fill
            row_fill = None
            if add_row_colors:
                if row_idx % 2 == 0:
                    row_fill = PatternFill(
                        start_color=ColorScheme.ACCENT_GRAY,
                        end_color=ColorScheme.ACCENT_GRAY,
                        fill_type="solid"
                    )

            for col_idx, (field, _, data_type, _) in enumerate(columns, 1):
                value = self._safe_get_value(item, field)
                cell = ws.cell(row=row_idx, column=col_idx)

                # Apply cell formatting based on data type
                if data_type == "integer":
                    cell.value = self._safe_int(value)
                    cell.number_format = ExcelStyles.INTEGER_FORMAT
                    cell.alignment = ExcelStyles.ALIGN_RIGHT
                elif data_type == "number":
                    cell.value = self._safe_float(value)
                    cell.number_format = ExcelStyles.NUMBER_FORMAT
                    cell.alignment = ExcelStyles.ALIGN_RIGHT
                elif data_type == "percent":
                    # 特殊处理百分比列
                    if field in percent_columns:
                        cell.value = self._safe_float(value)
                    else:
                        cell.value = self._safe_float(value) / 100
                    cell.number_format = ExcelStyles.PERCENT_FORMAT
                    cell.alignment = ExcelStyles.ALIGN_RIGHT
                elif data_type == "date":
                    cell.value = self._parse_datetime(value)
                    cell.number_format = ExcelStyles.DATE_FORMAT
                    cell.alignment = ExcelStyles.ALIGN_CENTER
                else:
                    cell.value = str(value) if value is not None else ""
                    cell.alignment = ExcelStyles.ALIGN_LEFT

                cell.border = ExcelStyles.THIN_BORDER

                # Apply row color
                if row_fill:
                    cell.fill = row_fill

                # Apply custom color based on column value (overrides row color)
                if color_column and field == color_column and color_mapper:
                    cell.fill = color_mapper(value)

    def _write_section_header(
        self,
        ws,
        row: int,
        text: str,
        start_col: int = 1,
        end_col: int = 2,
    ) -> None:
        """Write a section header with formatting."""
        ws.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
        cell = ws.cell(row=row, column=start_col)
        cell.value = text
        cell.font = ExcelStyles.SUBTITLE_FONT
        cell.fill = ExcelStyles.SUBTITLE_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')

    def _write_info_row(
        self,
        ws,
        row: int,
        label1: str,
        value1: str,
        label2: str = "",
        value2: str = "",
    ) -> None:
        """Write an information row with two columns."""
        label_style = Font(bold=True, color=ColorScheme.PRIMARY_DARK)

        ws.cell(row=row, column=1, value=label1).font = label_style
        ws.cell(row=row, column=2, value=value1).alignment = ExcelStyles.ALIGN_LEFT

        if label2:
            ws.cell(row=row, column=3, value=label2).font = label_style
            ws.cell(row=row, column=4, value=value2).alignment = ExcelStyles.ALIGN_LEFT

    def _write_no_data(self, ws, row: int = 1, message: str = "无数据") -> None:
        """Write a no data message."""
        cell = ws.cell(row=row, column=1)
        cell.value = message
        cell.font = Font(size=12, italic=True, color="999999")
        cell.alignment = ExcelStyles.ALIGN_CENTER
        ws.merge_cells(f"A{row}:{get_column_letter(10)}{row}")

    def _format_datetime(self, dt_str: Optional[str]) -> str:
        """Format datetime string."""
        if not dt_str:
            return "-"

        try:
            dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(dt_str)

    def _parse_datetime(self, value: Any) -> Any:
        """Parse datetime for Excel cell."""
        if not value:
            return None

        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value.replace("Z", "+00:00"))
            except Exception:
                return value

        return value

    def _format_status(self, status: str) -> str:
        """Format task status."""
        status_map = {
            "pending": "待执行",
            "running": "执行中",
            "completed": "已完成",
            "failed": "失败",
            "cancelled": "已取消",
        }
        return status_map.get(status, status)

    def _format_grade(self, grade: str) -> str:
        """Format health grade."""
        grade_map = {
            "excellent": "优秀",
            "good": "良好",
            "fair": "一般",
            "poor": "较差",
            "critical": "危急",
            "unknown": "未知",
        }
        return grade_map.get(grade, grade)

    def _get_score_grade(self, score: float) -> str:
        """Get grade text from score."""
        if score >= 90:
            return "优秀"
        elif score >= 75:
            return "良好"
        elif score >= 60:
            return "一般"
        elif score >= 40:
            return "较差"
        else:
            return "危急"

    def _format_adjustment_type(self, adj_type: Optional[str]) -> str:
        """Format adjustment type."""
        if not adj_type:
            return "-"
        type_map = {
            "down_significant": "大幅缩减",
            "down": "缩减",
            "none": "合理",
            "up": "扩容",
            "up_significant": "大幅扩容",
        }
        return type_map.get(adj_type, adj_type)

    def _format_volatility_level(self, volatility: str) -> str:
        """Format volatility level."""
        volatility_map = {
            "low": "低",
            "moderate": "中",
            "high": "高",
            "unknown": "未知",
        }
        return volatility_map.get(volatility, volatility)

    def _safe_get_value(self, data: Dict, field: str) -> Any:
        """Safely get value from dict, handling None and missing keys."""
        if data is None:
            return None
        return data.get(field)

    def _safe_float(self, value: Any) -> float:
        """Safely convert value to float."""
        try:
            return float(value) if value is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _safe_int(self, value: Any) -> int:
        """Safely convert value to int."""
        try:
            return int(value) if value is not None else 0
        except (ValueError, TypeError):
            return 0

    def _auto_fit_columns(self, ws) -> None:
        """Auto-fit column widths with improved algorithm."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        # 计算中文字符（占2个英文字符宽度）
                        value_str = str(cell.value)
                        length = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in value_str)
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass

            # 设置合适的宽度，限制最大宽度
            adjusted_width = min(max_length + 2, 60)
            # 确保最小宽度
            adjusted_width = max(adjusted_width, 10)
            ws.column_dimensions[column_letter].width = adjusted_width
