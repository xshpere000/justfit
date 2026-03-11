"""Integration tests for Excel report generation.

This test uses mock data to generate a complete Excel report,
verifying that all components work together correctly.
"""

import sys
import os
from pathlib import Path

# Add backend directory to Python path BEFORE any other imports
project_root = Path(__file__).parent.parent.parent.parent
backend_dir = project_root / "backend"
sys.path.insert(0, str(backend_dir))

# Also ensure project root is in path for potential resource access
sys.path.insert(0, str(project_root))

# Set working directory to project root for relative path resolution
os.chdir(project_root)

import pytest
from datetime import datetime, timezone
from typing import Dict, Any, List

# Import openpyxl for verification
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def get_mock_report_data() -> Dict[str, Any]:
    """Generate comprehensive mock data for Excel report testing.

    Returns:
        Complete report data structure matching ReportBuilder output
    """
    return {
        "task": {
            "id": 1,
            "name": "测试评估任务",
            "type": "collection",
            "status": "completed",
            "created_at": "2026-03-09T10:00:00",
            "completed_at": "2026-03-09T12:00:00",
        },
        "connection": {
            "name": "测试-vCenter",
            "platform": "vmware",
        },
        "summary": {
            "total_clusters": 3,
            "total_hosts": 12,
            "total_vms": 156,
            "powered_on_vms": 142,
            "powered_off_vms": 14,
            "total_cpu_mhz": 480000,
            "total_memory_gb": 3840.0,
        },
        "resources": {
            "clusters": [
                {
                    "name": "Cluster-Production",
                    "datacenter": "DC1",
                    "total_cpu": 192000,
                    "total_memory_gb": 1536.0,
                    "num_hosts": 5,
                    "num_vms": 65,
                },
                {
                    "name": "Cluster-Development",
                    "datacenter": "DC1",
                    "total_cpu": 144000,
                    "total_memory_gb": 1152.0,
                    "num_hosts": 4,
                    "num_vms": 48,
                },
                {
                    "name": "Cluster-Testing",
                    "datacenter": "DC2",
                    "total_cpu": 144000,
                    "total_memory_gb": 1152.0,
                    "num_hosts": 3,
                    "num_vms": 43,
                },
            ],
            "hosts": [
                {
                    "name": "esxi-prod-01.example.com",
                    "datacenter": "DC1",
                    "ip_address": "10.0.1.101",
                    "cpu_cores": 24,
                    "cpu_mhz": 2400,
                    "memory_gb": 256.0,
                    "num_vms": 15,
                    "power_state": "poweredon",
                    "overall_status": "green",
                },
                {
                    "name": "esxi-prod-02.example.com",
                    "datacenter": "DC1",
                    "ip_address": "10.0.1.102",
                    "cpu_cores": 24,
                    "cpu_mhz": 2400,
                    "memory_gb": 256.0,
                    "num_vms": 18,
                    "power_state": "poweredon",
                    "overall_status": "green",
                },
                {
                    "name": "esxi-prod-03.example.com",
                    "datacenter": "DC1",
                    "ip_address": "10.0.1.103",
                    "cpu_cores": 24,
                    "cpu_mhz": 2400,
                    "memory_gb": 256.0,
                    "num_vms": 12,
                    "power_state": "poweredon",
                    "overall_status": "yellow",
                },
            ],
            "vms": [
                {
                    "name": "web-server-prod-01",
                    "datacenter": "DC1",
                    "cpu_count": 4,
                    "memory_gb": 16.0,
                    "power_state": "poweredon",
                    "guest_os": "Ubuntu Linux",
                    "ip_address": "10.0.10.51",
                    "host_ip": "10.0.1.101",
                    "connection_state": "connected",
                    "overall_status": "green",
                },
                {
                    "name": "db-server-prod-01",
                    "datacenter": "DC1",
                    "cpu_count": 8,
                    "memory_gb": 32.0,
                    "power_state": "poweredon",
                    "guest_os": "CentOS Linux",
                    "ip_address": "10.0.10.61",
                    "host_ip": "10.0.1.102",
                    "connection_state": "connected",
                    "overall_status": "green",
                },
                {
                    "name": "app-server-prod-01",
                    "datacenter": "DC1",
                    "cpu_count": 4,
                    "memory_gb": 8.0,
                    "power_state": "poweredon",
                    "guest_os": "Windows Server",
                    "ip_address": "10.0.10.71",
                    "host_ip": "10.0.1.101",
                    "connection_state": "connected",
                    "overall_status": "green",
                },
            ],
        },
        "analysis": {
            "health": {
                "overallScore": 78.5,
                "grade": "good",
                "balanceScore": 82.0,
                "overcommitScore": 75.0,
                "hotspotScore": 68.0,
                "clusterCount": 3,
                "hostCount": 12,
                "vmCount": 156,
                "avgVmDensity": 13.0,
                "findings": [
                    {
                        "severity": "medium",
                        "category": "balance",
                        "cluster": "Cluster-Production",
                        "description": "Cluster-Production 中主机间VM分布略显不均，变异系数为0.45",
                        "details": {},
                    },
                    {
                        "severity": "low",
                        "category": "overcommit",
                        "cluster": "Cluster-Development",
                        "description": "Cluster-Development 的CPU超配比为1.8，处于合理范围内",
                        "details": {},
                    },
                ],
                "recommendations": [
                    "建议对 Cluster-Production 中的VM进行轻微调度，以改善负载均衡",
                    "关注 esxi-prod-03 的资源使用情况，其整体状态为黄色",
                ],
            },
            "idle": [
                {
                    "vmName": "test-server-legacy-01",
                    "cluster": "Cluster-Development",
                    "hostIp": "10.0.1.104",
                    "idleType": "powered_off",
                    "confidence": 95.0,
                    "riskLevel": "critical",
                    "daysInactive": 45,
                    "cpuUsageP95": 0.0,
                    "memoryUsageP95": 0.0,
                    "cpuCores": 4,
                    "memoryGb": 16.0,
                    "recommendation": "该虚拟机已关机45天，建议删除或归档以释放资源",
                },
                {
                    "vmName": "backup-server-temp-02",
                    "cluster": "Cluster-Testing",
                    "hostIp": "10.0.1.108",
                    "idleType": "idle_powered_on",
                    "confidence": 88.0,
                    "riskLevel": "high",
                    "daysInactive": 21,
                    "cpuUsageP95": 3.2,
                    "memoryUsageP95": 12.5,
                    "cpuCores": 2,
                    "memoryGb": 8.0,
                    "recommendation": "该虚拟机CPU和内存使用率极低，建议关机或迁移到低配主机",
                },
                {
                    "vmName": "dev-server-idle-03",
                    "cluster": "Cluster-Development",
                    "hostIp": "10.0.1.105",
                    "idleType": "low_activity",
                    "confidence": 72.0,
                    "riskLevel": "medium",
                    "daysInactive": 14,
                    "cpuUsageP95": 8.5,
                    "memoryUsageP95": 18.3,
                    "cpuCores": 4,
                    "memoryGb": 16.0,
                    "recommendation": "该虚拟机活跃度较低，建议检查是否仍需运行",
                },
                {
                    # 测试关机VM（没有cpuCores和memoryGb的情况）
                    "vmName": "old-server-removed-04",
                    "cluster": "Cluster-Testing",
                    "hostIp": "",
                    "idleType": "powered_off",
                    "confidence": 98.0,
                    "riskLevel": "critical",
                    "recommendation": "该虚拟机长期关机，建议删除",
                },
            ],
            "resource": {
                "rightSize": [
                    {
                        "vmName": "web-server-overprovisioned",
                        "cluster": "Cluster-Production",
                        "hostIp": "10.0.1.101",
                        "currentConfig": {"cpu": 8, "memory": 32},
                        "recommendedConfig": {"cpu": 4, "memory": 16},
                        "wasteRatio": 50.0,
                        "adjustmentType": "downsize_both",
                        "recommendation": "建议大幅缩容至 4 vCPU / 16GB RAM，可节省50%资源",
                    },
                    {
                        "vmName": "app-server-underprovisioned",
                        "cluster": "Cluster-Production",
                        "hostIp": "10.0.1.102",
                        "currentConfig": {"cpu": 2, "memory": 4},
                        "recommendedConfig": {"cpu": 4, "memory": 8},
                        "wasteRatio": 0,
                        "adjustmentType": "upsize_both",
                        "recommendation": "CPU和内存使用率持续偏高，建议扩容至 4 vCPU / 8GB RAM",
                    },
                    {
                        "vmName": "batch-server-overallocated",
                        "cluster": "Cluster-Development",
                        "hostIp": "10.0.1.104",
                        "currentConfig": {"cpu": 16, "memory": 64},
                        "recommendedConfig": {"cpu": 8, "memory": 32},
                        "wasteRatio": 48.5,
                        "adjustmentType": "downsize",
                        "recommendation": "建议缩容至 8 vCPU / 32GB RAM",
                    },
                ],
                "usagePattern": [
                    {
                        "vmName": "report-server-tidal",
                        "datacenter": "DC1",
                        "cluster": "Cluster-Production",
                        "hostIp": "10.0.1.101",
                        "optimizationType": "usage_pattern",
                        "usagePattern": "tidal",
                        "volatilityLevel": "high",
                        "coefficientOfVariation": 0.72,
                        "peakValleyRatio": 4.5,
                        "tidalDetails": {
                            "patternType": "day_active",
                            "dayAvg": 65.3,
                            "nightAvg": 12.8,
                        },
                        "recommendation": "呈现明显的潮汐模式，白天活跃夜间空闲，建议配置自动调度",
                    },
                    {
                        "vmName": "batch-server-burst",
                        "datacenter": "DC1",
                        "cluster": "Cluster-Development",
                        "hostIp": "10.0.1.104",
                        "optimizationType": "usage_pattern",
                        "usagePattern": "burst",
                        "volatilityLevel": "high",
                        "coefficientOfVariation": 0.85,
                        "peakValleyRatio": 6.2,
                        "recommendation": "呈现突发使用模式，建议配置自动扩缩容策略",
                    },
                    {
                        "vmName": "core-server-stable",
                        "datacenter": "DC1",
                        "cluster": "Cluster-Production",
                        "hostIp": "10.0.1.102",
                        "optimizationType": "usage_pattern",
                        "usagePattern": "stable",
                        "volatilityLevel": "low",
                        "coefficientOfVariation": 0.15,
                        "peakValleyRatio": 1.3,
                        "recommendation": "使用率稳定，当前配置合理",
                    },
                ],
                "mismatch": [
                    {
                        "vmName": "server-cpu-rich-memory-poor",
                        "datacenter": "DC1",
                        "cluster": "Cluster-Production",
                        "hostIp": "10.0.1.101",
                        "optimizationType": "resource_mismatch",
                        "hasMismatch": True,
                        "mismatchType": "cpu_rich_memory_poor",
                        "cpuUtilization": 15.2,
                        "memoryUtilization": 82.5,
                        "currentCpu": 8,
                        "currentMemory": 16.0,
                        "recommendation": "CPU使用率低但内存使用率高，建议降低CPU配置或增加内存",
                    },
                    {
                        "vmName": "server-cpu-poor-memory-rich",
                        "datacenter": "DC1",
                        "cluster": "Cluster-Development",
                        "hostIp": "10.0.1.104",
                        "optimizationType": "resource_mismatch",
                        "hasMismatch": True,
                        "mismatchType": "cpu_poor_memory_rich",
                        "cpuUtilization": 78.3,
                        "memoryUtilization": 22.8,
                        "currentCpu": 4,
                        "currentMemory": 32.0,
                        "recommendation": "CPU使用率高但内存使用率低，建议增加CPU配置或降低内存",
                    },
                    {
                        "vmName": "server-both-underutilized",
                        "datacenter": "DC2",
                        "cluster": "Cluster-Testing",
                        "hostIp": "10.0.1.108",
                        "optimizationType": "resource_mismatch",
                        "hasMismatch": True,
                        "mismatchType": "both_underutilized",
                        "cpuUtilization": 18.5,
                        "memoryUtilization": 25.3,
                        "currentCpu": 8,
                        "currentMemory": 32.0,
                        "recommendation": "CPU和内存使用率都偏低，资源被过度分配",
                    },
                ],
                "summary": {
                    "rightSizeCount": 3,
                    "usagePatternCount": 3,
                    "mismatchCount": 3,
                    "totalVmsAnalyzed": 156,
                },
            },
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@pytest.mark.asyncio
async def test_excel_report_generation_full(tmp_path):
    """Test complete Excel report generation with mock data.

    This test verifies that the ExcelReportGenerator can handle
    all types of data and generate a valid Excel file.

    Args:
        tmp_path: Pytest temporary path fixture
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    # Import here to avoid import errors if openpyxl is not installed
    from app.report.excel import ExcelReportGenerator

    # Get mock data
    data = get_mock_report_data()

    # Create output path
    output_path = tmp_path / "test_report.xlsx"

    # Generate Excel
    generator = ExcelReportGenerator()
    result_path = generator.generate(data, str(output_path))

    # Verify Excel file was created
    assert Path(result_path).exists()
    assert Path(result_path).stat().st_size > 0

    # Verify file extension
    assert result_path.endswith(".xlsx")

    # Verify file is reasonably sized
    file_size = Path(result_path).stat().st_size
    assert file_size > 5000, f"Excel file too small: {file_size} bytes"

    # Load and verify workbook structure
    wb = load_workbook(result_path)
    sheet_names = wb.sheetnames

    # Verify all expected sheets exist
    expected_sheets = [
        "概览",
        "集群",
        "主机",
        "虚拟机",
        "闲置检测",
        "Right Size优化",
        "使用模式分析",
        "配置错配分析",
        "健康评分",
    ]
    for expected in expected_sheets:
        assert expected in sheet_names, f"Missing sheet: {expected}"

    # Verify summary sheet has content
    summary_ws = wb["概览"]
    assert summary_ws["A1"].value == "云平台资源评估报告"

    # Verify idle detection sheet has data with risk level colors
    idle_ws = wb["闲置检测"]
    # Check header row
    assert idle_ws["A1"].value == "虚拟机名称"
    assert idle_ws["I1"].value == "建议"
    # Check we have data rows (4 idle VMs + header = 5 rows)
    assert idle_ws.max_row >= 5

    # Verify risk level column has Chinese values
    risk_levels = [idle_ws.cell(row=i, column=8).value for i in range(2, min(6, idle_ws.max_row + 1))]
    assert any("危急" in str(rl) or "高" in str(rl) or "中" in str(rl) for rl in risk_levels if rl)

    wb.close()


@pytest.mark.asyncio
async def test_excel_report_generation_minimal(tmp_path):
    """Test Excel report generation with minimal data.

    This test verifies edge cases like empty analysis results.

    Args:
        tmp_path: Pytest temporary path fixture
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    from app.report.excel import ExcelReportGenerator

    # Minimal data with empty analysis results
    data = {
        "task": {
            "id": 1,
            "name": "最小测试任务",
            "type": "collection",
            "status": "completed",
            "created_at": "2026-03-09T10:00:00",
            "completed_at": "2026-03-09T12:00:00",
        },
        "connection": {
            "name": "测试平台",
            "platform": "vmware",
        },
        "summary": {
            "total_clusters": 0,
            "total_hosts": 0,
            "total_vms": 0,
            "powered_on_vms": 0,
            "powered_off_vms": 0,
            "total_cpu_mhz": 0,
            "total_memory_gb": 0.0,
        },
        "resources": {
            "clusters": [],
            "hosts": [],
            "vms": [],
        },
        "analysis": {
            "health": None,
            "idle": [],
            "resource": {
                "rightSize": [],
                "usagePattern": [],
                "mismatch": [],
            },
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    output_path = tmp_path / "test_minimal_report.xlsx"

    generator = ExcelReportGenerator()
    result_path = generator.generate(data, str(output_path))

    assert Path(result_path).exists()
    assert Path(result_path).stat().st_size > 0

    # Load and verify empty data handling
    wb = load_workbook(result_path)

    # Check sheets with no data show "无数据" message
    clusters_ws = wb["集群"]
    assert clusters_ws["A1"].value == "无数据"

    idle_ws = wb["闲置检测"]
    assert idle_ws["A1"].value == "无数据"

    health_ws = wb["健康评分"]
    assert health_ws["A1"].value == "无健康评分数据"

    wb.close()


@pytest.mark.asyncio
async def test_excel_idle_sheet_with_missing_fields(tmp_path):
    """Test idle sheet handling of VMs with missing cpuCores/memoryGb.

    This verifies that powered-off VMs (which don't have these fields)
    are handled correctly.

    Args:
        tmp_path: Pytest temporary path fixture
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    from app.report.excel import ExcelReportGenerator

    # Data with powered-off VM that has no cpuCores/memoryGb
    data = {
        "task": {
            "id": 1,
            "name": "测试任务",
            "type": "collection",
            "status": "completed",
            "created_at": "2026-03-09T10:00:00",
            "completed_at": "2026-03-09T12:00:00",
        },
        "connection": {
            "name": "测试平台",
            "platform": "vmware",
        },
        "summary": {
            "total_clusters": 1,
            "total_hosts": 1,
            "total_vms": 2,
            "powered_on_vms": 1,
            "powered_off_vms": 1,
            "total_cpu_mhz": 48000,
            "total_memory_gb": 384.0,
        },
        "resources": {
            "clusters": [],
            "hosts": [],
            "vms": [],
        },
        "analysis": {
            "idle": [
                {
                    # Powered-on VM with all fields
                    "vmName": "active-vm-01",
                    "cluster": "Cluster-01",
                    "hostIp": "10.0.1.101",
                    "cpuCores": 4,
                    "memoryGb": 16.0,
                    "idleType": "idle_powered_on",
                    "confidence": 85.0,
                    "riskLevel": "high",
                    "recommendation": "建议关机",
                },
                {
                    # Powered-off VM without cpuCores/memoryGb
                    "vmName": "powered-off-vm-02",
                    "cluster": "Cluster-01",
                    "hostIp": "",
                    "idleType": "powered_off",
                    "confidence": 95.0,
                    "riskLevel": "critical",
                    "recommendation": "建议删除",
                },
            ],
            "resource": {
                "rightSize": [],
                "usagePattern": [],
                "mismatch": [],
            },
            "health": None,
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    output_path = tmp_path / "test_missing_fields.xlsx"

    generator = ExcelReportGenerator()
    result_path = generator.generate(data, str(output_path))

    # Load and verify
    wb = load_workbook(result_path)
    idle_ws = wb["闲置检测"]

    # Check that both rows exist (header + 2 data rows)
    assert idle_ws.max_row == 3

    # Check first row (has all fields)
    assert idle_ws["A2"].value == "active-vm-01"
    assert idle_ws["D2"].value == 4  # cpuCores
    assert idle_ws["E2"].value == 16.0  # memoryGb

    # Check second row (powered-off, missing fields should be 0)
    assert idle_ws["A3"].value == "powered-off-vm-02"
    assert idle_ws["D3"].value == 0  # cpuCores defaults to 0
    assert idle_ws["E3"].value == 0  # memoryGb defaults to 0

    wb.close()


@pytest.mark.asyncio
async def test_excel_risk_level_color_mapping(tmp_path):
    """Test that risk levels are correctly mapped to colors.

    Args:
        tmp_path: Pytest temporary path fixture
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    from app.report.excel import ExcelReportGenerator, RiskColors

    # Test color mapping directly
    test_cases = [
        ("critical", "危急"),
        ("high", "高"),
        ("medium", "中"),
        ("low", "低"),
    ]

    for english, chinese in test_cases:
        fill_en = RiskColors.get_fill(english)
        fill_cn = RiskColors.get_fill(chinese)

        # Both should have the same color
        assert fill_en.start_color.rgb == fill_cn.start_color.rgb, \
            f"Color mismatch for {english}/{chinese}"

    # Now test in actual Excel generation
    data = {
        "task": {
            "id": 1,
            "name": "颜色测试",
            "type": "collection",
            "status": "completed",
            "created_at": "2026-03-09T10:00:00",
            "completed_at": "2026-03-09T12:00:00",
        },
        "connection": {
            "name": "测试",
            "platform": "vmware",
        },
        "summary": {
            "total_clusters": 1,
            "total_hosts": 1,
            "total_vms": 4,
            "powered_on_vms": 4,
            "powered_off_vms": 0,
            "total_cpu_mhz": 48000,
            "total_memory_gb": 384.0,
        },
        "resources": {
            "clusters": [],
            "hosts": [],
            "vms": [],
        },
        "analysis": {
            "idle": [
                {"vmName": f"vm-{rl}", "cluster": "C1", "hostIp": "10.0.0.1",
                 "cpuCores": 2, "memoryGb": 8, "idleType": "idle_powered_on",
                 "confidence": 80, "riskLevel": rl, "recommendation": "测试"}
                for rl in ["critical", "high", "medium", "low"]
            ],
            "resource": {"rightSize": [], "usagePattern": [], "mismatch": []},
            "health": None,
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    output_path = tmp_path / "test_colors.xlsx"
    generator = ExcelReportGenerator()
    result_path = generator.generate(data, str(output_path))

    # Verify colors were applied
    wb = load_workbook(result_path)
    idle_ws = wb["闲置检测"]

    # Get fill colors for each risk level row
    expected_colors = {
        "critical": "00FF0000",  # Red
        "高": "FFFF6B6B",        # Light red
        "中": "00FFC000",        # Yellow/Orange
        "低": "0070AD47",        # Green
    }

    for i, (rl, expected_color) in enumerate(expected_colors.items(), 2):
        cell = idle_ws.cell(row=i, column=8)  # riskLevel column
        fill_color = cell.fill.start_color.rgb if cell.fill.start_color else None
        assert fill_color == expected_color, \
            f"Row {i} ({rl}): expected {expected_color}, got {fill_color}"

    wb.close()


@pytest.mark.asyncio
async def test_excel_health_score_colors(tmp_path):
    """Test that health scores are correctly colored.

    Args:
        tmp_path: Pytest temporary path fixture
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    from app.report.excel import ExcelReportGenerator, HealthColors, ColorScheme

    # Test color mapping
    assert HealthColors.get_color(90) == ColorScheme.SUCCESS  # Green
    assert HealthColors.get_color(70) == ColorScheme.WARNING  # Yellow
    assert HealthColors.get_color(50) == "#FF9900"           # Orange
    assert HealthColors.get_color(30) == ColorScheme.DANGER   # Red

    # Test with actual data
    data = {
        "task": {
            "id": 1,
            "name": "健康评分测试",
            "type": "collection",
            "status": "completed",
            "created_at": "2026-03-09T10:00:00",
            "completed_at": "2026-03-09T12:00:00",
        },
        "connection": {
            "name": "测试",
            "platform": "vmware",
        },
        "summary": {
            "total_clusters": 1,
            "total_hosts": 1,
            "total_vms": 10,
            "powered_on_vms": 10,
            "powered_off_vms": 0,
            "total_cpu_mhz": 48000,
            "total_memory_gb": 384.0,
        },
        "resources": {
            "clusters": [],
            "hosts": [],
            "vms": [],
        },
        "analysis": {
            "idle": [],
            "resource": {"rightSize": [], "usagePattern": [], "mismatch": []},
            "health": {
                "overallScore": 75,
                "grade": "good",
                "balanceScore": 82,
                "overcommitScore": 70,
                "hotspotScore": 68,
                "clusterCount": 1,
                "hostCount": 1,
                "vmCount": 10,
                "findings": [],
            },
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    output_path = tmp_path / "test_health.xlsx"
    generator = ExcelReportGenerator()
    result_path = generator.generate(data, str(output_path))

    # Verify health sheet has colored scores
    wb = load_workbook(result_path)
    health_ws = wb["健康评分"]

    # Check overall score cell color (75 -> yellow)
    score_cell = health_ws["B2"]
    score_color = score_cell.font.color.rgb if score_cell.font.color else None
    assert score_color == "00FFC000", f"Overall score color wrong: {score_color}"

    wb.close()


@pytest.mark.asyncio
async def test_excel_percentage_formatting(tmp_path):
    """Test that percentage fields are correctly formatted.

    Args:
        tmp_path: Pytest temporary path fixture
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    from app.report.excel import ExcelReportGenerator

    data = {
        "task": {
            "id": 1,
            "name": "百分比格式测试",
            "type": "collection",
            "status": "completed",
            "created_at": "2026-03-09T10:00:00",
            "completed_at": "2026-03-09T12:00:00",
        },
        "connection": {
            "name": "测试",
            "platform": "vmware",
        },
        "summary": {
            "total_clusters": 1,
            "total_hosts": 1,
            "total_vms": 3,
            "powered_on_vms": 3,
            "powered_off_vms": 0,
            "total_cpu_mhz": 48000,
            "total_memory_gb": 384.0,
        },
        "resources": {
            "clusters": [],
            "hosts": [],
            "vms": [],
        },
        "analysis": {
            "idle": [],
            "resource": {
                "rightSize": [
                    {
                        "vmName": f"vm-{i}",
                        "cluster": "C1",
                        "currentConfig": {"cpu": 8, "memory": 32},
                        "recommendedConfig": {"cpu": 4, "memory": 16},
                        "wasteRatio": 50.0 * i,  # 0%, 50%, 100%
                        "adjustmentType": "downsize",
                        "recommendation": f"测试{i}",
                    }
                    for i in range(1, 4)
                ],
                "usagePattern": [],
                "mismatch": [],
            },
            "health": None,
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    output_path = tmp_path / "test_percentages.xlsx"
    generator = ExcelReportGenerator()
    result_path = generator.generate(data, str(output_path))

    # Verify percentage formatting
    wb = load_workbook(result_path)
    rightsize_ws = wb["Right Size优化"]

    # Check wasteRatio column (column G) has percentage format
    for row in range(2, 5):
        cell = rightsize_ws.cell(row=row, column=7)  # wasteRatio column
        assert cell.number_format == "0.0%", \
            f"Row {row} wasteRatio format: {cell.number_format}"

    wb.close()


@pytest.mark.asyncio
async def test_excel_chinese_mappings(tmp_path):
    """Test that all Chinese mappings work correctly.

    Args:
        tmp_path: Pytest temporary path fixture
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    from app.report.excel import ExcelReportGenerator

    data = {
        "task": {
            "id": 1,
            "name": "中文映射测试",
            "type": "collection",
            "status": "completed",
            "created_at": "2026-03-09T10:00:00",
            "completed_at": "2026-03-09T12:00:00",
        },
        "connection": {
            "name": "测试",
            "platform": "vmware",
        },
        "summary": {
            "total_clusters": 1,
            "total_hosts": 1,
            "total_vms": 10,
            "powered_on_vms": 10,
            "powered_off_vms": 0,
            "total_cpu_mhz": 48000,
            "total_memory_gb": 384.0,
        },
        "resources": {
            "clusters": [],
            "hosts": [],
            "vms": [],
        },
        "analysis": {
            "idle": [
                {
                    "vmName": f"vm-{t}",
                    "cluster": "C1",
                    "hostIp": "10.0.0.1",
                    "cpuCores": 2,
                    "memoryGb": 8,
                    "idleType": t,
                    "confidence": 80,
                    "riskLevel": "low",
                    "recommendation": "测试",
                }
                for t in ["powered_off", "idle_powered_on", "low_activity"]
            ],
            "resource": {
                "rightSize": [
                    {
                        "vmName": f"vm-{t}",
                        "cluster": "C1",
                        "currentConfig": {"cpu": 4, "memory": 16},
                        "recommendedConfig": {"cpu": 2, "memory": 8},
                        "wasteRatio": 50,
                        "adjustmentType": t,
                        "recommendation": "测试",
                    }
                    for t in ["downsize_both", "upsize_both", "downsize_cpu"]
                ],
                "usagePattern": [
                    {
                        "vmName": f"vm-{p}",
                        "cluster": "C1",
                        "usagePattern": p,
                        "volatilityLevel": "low",
                        "coefficientOfVariation": 0.2,
                        "peakValleyRatio": 1.5,
                        "recommendation": "测试",
                    }
                    for p in ["stable", "burst", "tidal"]
                ],
                "mismatch": [
                    {
                        "vmName": f"vm-{m}",
                        "cluster": "C1",
                        "mismatchType": m,
                        "cpuUtilization": 20,
                        "memoryUtilization": 80,
                        "currentCpu": 4,
                        "currentMemory": 16,
                        "recommendation": "测试",
                    }
                    for m in ["cpu_rich_memory_poor", "cpu_poor_memory_rich", "both_underutilized"]
                ],
            },
            "health": None,
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    output_path = tmp_path / "test_mappings.xlsx"
    generator = ExcelReportGenerator()
    result_path = generator.generate(data, str(output_path))

    # Verify Chinese mappings
    wb = load_workbook(result_path)

    # Check idle type mappings
    idle_ws = wb["闲置检测"]
    idle_types = [idle_ws.cell(row=i, column=6).value for i in range(2, 5)]
    assert "已关机" in idle_types
    assert "开机闲置" in idle_types
    assert "低活跃" in idle_types

    # Check adjustment type mappings
    rightsize_ws = wb["Right Size优化"]
    adj_types = [rightsize_ws.cell(row=i, column=8).value for i in range(2, 5)]
    assert "降配置" in adj_types
    assert "升配置" in adj_types

    # Check usage pattern mappings
    pattern_ws = wb["使用模式分析"]
    patterns = [pattern_ws.cell(row=i, column=3).value for i in range(2, 5)]
    assert "稳定" in patterns
    assert "突发" in patterns
    assert "潮汐" in patterns

    # Check mismatch type mappings
    mismatch_ws = wb["配置错配分析"]
    mismatches = [mismatch_ws.cell(row=i, column=3).value for i in range(2, 5)]
    assert "CPU富足/内存不足" in mismatches
    assert "CPU不足/内存富足" in mismatches
    assert "均利用不足" in mismatches

    wb.close()


if __name__ == "__main__":
    """Run tests directly for development."""
    import asyncio
    from tempfile import TemporaryDirectory

    async def run_tests():
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)

            print("Testing full Excel generation...")
            await test_excel_report_generation_full(tmp_path)
            print("✓ Full Excel generation test passed")

            print("Testing minimal Excel generation...")
            await test_excel_report_generation_minimal(tmp_path)
            print("✓ Minimal Excel generation test passed")

            print("Testing missing fields handling...")
            await test_excel_idle_sheet_with_missing_fields(tmp_path)
            print("✓ Missing fields test passed")

            print("Testing risk level color mapping...")
            await test_excel_risk_level_color_mapping(tmp_path)
            print("✓ Risk level color mapping test passed")

            print("Testing health score colors...")
            await test_excel_health_score_colors(tmp_path)
            print("✓ Health score colors test passed")

            print("Testing percentage formatting...")
            await test_excel_percentage_formatting(tmp_path)
            print("✓ Percentage formatting test passed")

            print("Testing Chinese mappings...")
            await test_excel_chinese_mappings(tmp_path)
            print("✓ Chinese mappings test passed")

            print("\nAll tests passed! ✓")
            print(f"Test Excel files saved to: {tmp_path}")

    asyncio.run(run_tests())
